import numpy as np
from openpyxl import Workbook

class NumericalDifferentiationDataset:
    def __init__(self, x_values, y_values, h):
        self.x_values = np.array(x_values)
        self.y_values = np.array(y_values)
        self.h = h
        self.three_point_results = []
        self.five_point_results = []
        self.observations = []

    def three_point_formula(self):
        n = len(self.x_values)
        derivatives = np.zeros(n)
        self.observations = ["Ec. progresiva"] + ["Ec. centrada"] * (n - 2) + ["Ec. regresiva"]

        # Forward difference for the first point
        derivatives[0] = (-3 * self.y_values[0] + 
                          4 * self.y_values[1] - 
                          self.y_values[2]) / (2 * self.h)

        # Centered differences for interior points
        for i in range(1, n - 1):
            derivatives[i] = (self.y_values[i + 1] - 
                              self.y_values[i - 1]) / (2 * self.h)

        # Backward difference for the last point
        derivatives[n - 1] = (self.y_values[n - 3] - 
                              4 * self.y_values[n - 2] + 
                              3 * self.y_values[n - 1]) / (2 * self.h)

        self.three_point_results = derivatives
        return derivatives

    def five_point_formula(self):
        n = len(self.x_values)
        derivatives = np.zeros(n)

        # Forward difference for the first two points
        derivatives[0] = (-3 * self.y_values[0] +
                          4 * self.y_values[1] -
                          self.y_values[2]) / (2 * self.h)

        derivatives[1] = (self.y_values[2] -
                          self.y_values[0]) / (2 * self.h)

        # Centered differences for interior points
        for i in range(2, n - 2):
            derivatives[i] = (-self.y_values[i + 2] +
                              8 * self.y_values[i + 1] -
                              8 * self.y_values[i - 1] +
                              self.y_values[i - 2]) / (12 * self.h)

        # Backward difference for the last two points
        derivatives[n - 2] = (self.y_values[n - 1] -
                              self.y_values[n - 3]) / (2 * self.h)

        derivatives[n - 1] = (3 * self.y_values[n - 1] -
                              4 * self.y_values[n - 2] +
                              self.y_values[n - 3]) / (2 * self.h)

        self.five_point_results = derivatives
        return derivatives

    def save_to_excel(self, filename):
        wb = Workbook()

        # Three-point sheet
        ws_three = wb.create_sheet(title="Three-Point Formula")
        ws_three.append(["Xi", "Yi", "f'(Xi) (Calc)", "Observaciones"])

        for i in range(len(self.x_values)):
            ws_three.append([
                self.x_values[i],
                self.y_values[i],
                self.three_point_results[i],
                self.observations[i]
            ])

        # Five-point sheet
        ws_five = wb.create_sheet(title="Five-Point Formula")
        ws_five.append(["Xi", "Yi", "f'(Xi) (Calc)", "Observaciones"])

        for i in range(len(self.x_values)):
            ws_five.append([
                self.x_values[i],
                self.y_values[i],
                self.five_point_results[i],
                self.observations[i]
            ])

        # Save workbook
        wb.save(filename)

# Example usage with datasets
datasets = [
    ([0, 0.5, 1.0, 1.5, 2.0], [0, 0.4207, 0.4546, 0.0706, -0.3784], 0.5),  # Dataset 1.4
    ([0, 1, 2, 3, 4], [1, 1.1052, 1.2214, 1.3499, 1.4918], 1.0)  # Dataset 1.5
]

for i, (x_vals, y_vals, h) in enumerate(datasets, 1):
    print(f"\nProcessing Dataset {i}...")
    diff = NumericalDifferentiationDataset(x_vals, y_vals, h)
    three_point = diff.three_point_formula()
    five_point = diff.five_point_formula()
    filename = f"Numerical_Differentiation_Dataset_{i}.xlsx"
    diff.save_to_excel(filename)
    print(f"Results saved to {filename}.")
