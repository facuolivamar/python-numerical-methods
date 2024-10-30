import numpy as np
import openpyxl
from openpyxl import Workbook

class NumericalDifferentiation:
    def __init__(self, function, exact_derivative, x_values, h):
        self.function = function
        self.exact_derivative = exact_derivative
        self.x_values = np.array(x_values)
        self.h = h
        self.three_point_results = []
        self.five_point_results = []
        self.errors_three_point = []
        self.errors_five_point = []
        self.observations = []

    def three_point_formula(self):
        n = len(self.x_values)
        derivatives = np.zeros(n)
        self.observations = ["Ec. progresiva"] + ["Ec. centrada"] * (n - 2) + ["Ec. regresiva"]

        # Forward difference for the first point
        derivatives[0] = (-3 * self.function(self.x_values[0]) +
                          4 * self.function(self.x_values[1]) -
                          self.function(self.x_values[2])) / (2 * self.h)

        # Centered differences for interior points
        for i in range(1, n - 1):
            derivatives[i] = (self.function(self.x_values[i + 1]) -
                              self.function(self.x_values[i - 1])) / (2 * self.h)

        # Backward difference for the last point
        derivatives[n - 1] = (self.function(self.x_values[n - 3]) -
                              4 * self.function(self.x_values[n - 2]) +
                              3 * self.function(self.x_values[n - 1])) / (2 * self.h)

        self.three_point_results = derivatives
        self.errors_three_point = np.abs(derivatives - self.exact_derivative(self.x_values))
        return derivatives

    def five_point_formula(self):
        n = len(self.x_values)
        derivatives = np.zeros(n)

        # Forward and backward formulas at boundaries
        derivatives[0] = (-3 * self.function(self.x_values[0]) +
                          4 * self.function(self.x_values[1]) -
                          self.function(self.x_values[2])) / (2 * self.h)
        derivatives[1] = (self.function(self.x_values[2]) -
                          self.function(self.x_values[0])) / (2 * self.h)
        derivatives[n - 2] = (self.function(self.x_values[n - 1]) -
                              self.function(self.x_values[n - 3])) / (2 * self.h)
        derivatives[n - 1] = (3 * self.function(self.x_values[n - 1]) -
                              4 * self.function(self.x_values[n - 2]) +
                              self.function(self.x_values[n - 3])) / (2 * self.h)

        # Centered differences for the interior points
        for i in range(2, n - 2):
            derivatives[i] = (-self.function(self.x_values[i + 2]) +
                              8 * self.function(self.x_values[i + 1]) -
                              8 * self.function(self.x_values[i - 1]) +
                              self.function(self.x_values[i - 2])) / (12 * self.h)

        self.five_point_results = derivatives
        self.errors_five_point = np.abs(derivatives - self.exact_derivative(self.x_values))
        return derivatives

    def save_to_excel(self, filename):
        wb = Workbook()

        # Three-point sheet
        ws_three = wb.create_sheet(title="Three-Point Formula")
        ws_three.append(["Xi", "f(Xi)", "f'(Xi) (Calc)", "f'(Xi) (Exact)", "|E|", "Observaciones"])

        for i in range(len(self.x_values)):
            ws_three.append([
                self.x_values[i],
                self.function(self.x_values[i]),
                self.three_point_results[i],
                self.exact_derivative(self.x_values[i]),
                self.errors_three_point[i],
                self.observations[i]
            ])
            print(self.errors_three_point[i])

        # Five-point sheet
        ws_five = wb.create_sheet(title="Five-Point Formula")
        ws_five.append(["Xi", "f(Xi)", "f'(Xi) (Calc)", "f'(Xi) (Exact)", "|E|", "Observaciones"])

        for i in range(len(self.x_values)):
            ws_five.append([
                self.x_values[i],
                self.function(self.x_values[i]),
                self.five_point_results[i],
                self.exact_derivative(self.x_values[i]),
                self.errors_five_point[i],
                self.observations[i]
            ])

        # Save workbook
        wb.save(filename)

# Define the function and its exact derivative
def f1_1(x):
    return np.exp(x) * np.cos(x)

def f1_1_derivative(x):
    return np.exp(x) * (np.cos(x) - np.sin(x))

# Example usage
x_values = np.arange(0, 0.8, 0.1)  # [0, 0.1, 0.2, ..., 0.7]
h = 0.1

diff = NumericalDifferentiation(f1_1, f1_1_derivative, x_values, h)
diff.three_point_formula()
diff.five_point_formula()
diff.save_to_excel("Numerical_Differentiation_Results.xlsx")

print("Results saved to Numerical_Differentiation_Results.xlsx")
